"""
AI-Powered Statute Date Extractor and Filler

A comprehensive tool that combines Excel processing, AI-powered date extraction, and GUI-based review
for automatically filling missing dates in statute databases.

Features:
- Excel file processing with NumPy optimization
- GPT API integration for intelligent date extraction from statute sections
- Interactive GUI for reviewing and approving extracted dates
- Batch processing with progress tracking
- Configurable AI prompts and confidence scoring
- Export capabilities for processed data

Excel Column Structure:
- Column A: Statute_Name
- Column B: Best_Date (primary date to fill)
- Column C: All_Dates_Extracted (comma-separated if multiple)
- Column D: Selection_Reason
- Column E: Search_Method
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import numpy as np
import openpyxl
from pymongo import MongoClient
import json
import os
import sys
import re
from datetime import datetime
from typing import List, Dict, Optional, Tuple
import threading
from rapidfuzz import process, fuzz
from openai import AzureOpenAI
from dataclasses import dataclass
import calendar

# Add parent directory to path for utils imports
parent_dir = os.path.join(os.path.dirname(__file__), '..')
sys.path.append(parent_dir)
print(f"Added to path: {parent_dir}")

# Import GPT optimization utilities
try:
    from utils.gpt_cache import gpt_cache
    from utils.gpt_rate_limiter import AdvancedRateLimiter
    from utils.gpt_prompt_optimizer import PromptOptimizer
    from utils.gpt_monitor import gpt_monitor
    GPT_UTILS_AVAILABLE = True
    print("✅ GPT optimization utilities imported successfully")
except ImportError as e:
    GPT_UTILS_AVAILABLE = False
    print(f"⚠️ Warning: GPT optimization utilities not available: {e}")
    print(f"Current sys.path: {sys.path}")

def format_date_to_dd_mmm_yyyy(date_str: str) -> str:
    """Convert various date formats to DD-MMM-YYYY format"""
    if not date_str or not date_str.strip():
        return ""
    
    date_str = date_str.strip()
    
    # If already in DD-MMM-YYYY format, return as is
    if re.match(r'^\d{1,2}-[A-Za-z]{3}-\d{4}$', date_str):
        return date_str
    
    try:
        # Try different date formats
        date_formats = [
            '%Y-%m-%d',      # YYYY-MM-DD
            '%d/%m/%Y',      # DD/MM/YYYY
            '%m/%d/%Y',      # MM/DD/YYYY
            '%d-%m-%Y',      # DD-MM-YYYY
            '%Y/%m/%d',      # YYYY/MM/DD
            '%d-%m-%y',      # DD-MM-YY
            '%d/%m/%y',      # DD/MM/YY
            '%Y-%m-%d %H:%M:%S',  # YYYY-MM-DD HH:MM:SS
            '%d-%m-%Y %H:%M:%S',  # DD-MM-YYYY HH:MM:SS
            '%d-%m-%Y',      # DD-M-YYYY (e.g., 4-3-2016)
        ]
        
        parsed_date = None
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(date_str, fmt)
                break
            except ValueError:
                continue
        
        if parsed_date:
            # Format as DD-MMM-YYYY
            return parsed_date.strftime('%d-%b-%Y')
        else:
            # Try to handle ordinal dates and other complex formats
            # Handle ordinal dates like "4th March, 2016"
            ordinal_match = re.search(r'(\d{1,2})(st|nd|rd|th)\s+(\w+)\s*,\s*(\d{4})', date_str, re.IGNORECASE)
            if ordinal_match:
                day = ordinal_match.group(1)
                month_name = ordinal_match.group(3)
                year = ordinal_match.group(4)
                try:
                    # Convert month name to number
                    month_num = datetime.strptime(month_name, '%B').month
                    parsed_date = datetime(int(year), month_num, int(day))
                    return parsed_date.strftime('%d-%b-%Y')
                except:
                    pass
            
            # Handle "Month DD, YYYY" format like "March 3, 2016"
            month_day_match = re.search(r'(\w+)\s+(\d{1,2})\s*,\s*(\d{4})', date_str, re.IGNORECASE)
            if month_day_match:
                month_name = month_day_match.group(1)
                day = month_day_match.group(2)
                year = month_day_match.group(3)
                try:
                    # Convert month name to number
                    month_num = datetime.strptime(month_name, '%B').month
                    parsed_date = datetime(int(year), month_num, int(day))
                    return parsed_date.strftime('%d-%b-%Y')
                except:
                    pass
            
            # If all parsing fails, return original string
            return date_str
            
    except Exception as e:
        print(f"Error formatting date '{date_str}': {e}")
        return date_str

@dataclass
class StatuteRecord:
    """Data class for statute records"""
    excel_row: int
    statute_name: str
    best_date: str = ""
    all_dates_extracted: str = ""
    selection_reason: str = ""
    search_method: str = ""
    extracted_date: str = ""
    confidence: float = 0.0
    sections_used: List[str] = None
    
    def __post_init__(self):
        if self.sections_used is None:
            self.sections_used = []

class ExcelProcessor:
    """Handles Excel file operations"""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.headers = []
        
    def load_workbook(self):
        """Load Excel workbook and get headers"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            # Get headers from first row
            self.headers = []
            for cell in self.worksheet[1]:
                self.headers.append(str(cell.value).strip() if cell.value else "")
            
            print(f"Headers found: {self.headers}")
            return True
            
        except Exception as e:
            print(f"Error loading workbook: {e}")
            return False
    
    def get_statutes_with_missing_dates(self) -> List[StatuteRecord]:
        """Get statutes that have missing or incomplete Best_Date values"""
        statutes = []
        
        # Find column indices
        statute_name_col = self.get_column_index("Statute_Name")
        best_date_col = self.get_column_index("Best_Date")
        
        if statute_name_col is None or best_date_col is None:
            print("Error: Required columns not found")
            return statutes
        
        # Process each row
        for row_idx in range(2, self.worksheet.max_row + 1):  # Skip header row
            statute_name = self.get_cell_value(row_idx, statute_name_col)
            best_date = self.get_cell_value(row_idx, best_date_col)
            
            # Skip empty statute names
            if not statute_name or statute_name.strip() == "":
                continue
            
            # Check if Best_Date is missing, empty, or only contains year
            should_process = False
            
            if not best_date or best_date.strip() == "":
                # Missing or empty date
                should_process = True
            else:
                # Check if it's only a year (4 digits)
                import re
                year_pattern = r'^\d{4}$'
                if re.match(year_pattern, best_date.strip()):
                    # Only year found, need more specific date
                    should_process = True
                elif len(best_date.strip()) < 8:
                    # Very short date, likely incomplete
                    should_process = True
            
            if should_process:
                statute_record = StatuteRecord(
                    excel_row=row_idx,
                    statute_name=statute_name.strip()
                )
                statutes.append(statute_record)
        
        return statutes
    
    def get_column_index(self, column_name: str) -> Optional[int]:
        """Get column index by name"""
        try:
            return self.headers.index(column_name)
        except ValueError:
            return None
    
    def get_cell_value(self, row: int, col: int) -> str:
        """Get cell value as string"""
        try:
            cell_value = self.worksheet.cell(row=row, column=col + 1).value
            return str(cell_value) if cell_value is not None else ""
        except:
            return ""
    
    def save_to_excel(self, statute: StatuteRecord):
        """Save extracted data back to Excel"""
        try:
            # Column indices (0-based)
            best_date_col = self.get_column_index("Best_Date")
            all_dates_col = self.get_column_index("All_Dates_Extracted")
            selection_reason_col = self.get_column_index("Selection_Reason")
            search_method_col = self.get_column_index("Search_Method")
            
            # Format dates before saving
            formatted_best_date = format_date_to_dd_mmm_yyyy(statute.extracted_date)
            formatted_all_dates = format_date_to_dd_mmm_yyyy(statute.all_dates_extracted)
            
            # Update cells (openpyxl uses 1-based indexing)
            if best_date_col is not None:
                self.worksheet.cell(row=statute.excel_row, column=best_date_col + 1, value=formatted_best_date)
            
            if all_dates_col is not None:
                self.worksheet.cell(row=statute.excel_row, column=all_dates_col + 1, value=formatted_all_dates)
            
            if selection_reason_col is not None:
                self.worksheet.cell(row=statute.excel_row, column=selection_reason_col + 1, value=statute.selection_reason)
            
            if search_method_col is not None:
                self.worksheet.cell(row=statute.excel_row, column=search_method_col + 1, value=statute.search_method)
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")
    
    def save_workbook(self):
        """Save the workbook"""
        try:
            self.workbook.save(self.file_path)
            return True
        except Exception as e:
            print(f"Error saving workbook: {e}")
            return False

class DatabaseConnector:
    """Handles MongoDB database operations"""
    
    def __init__(self, mongo_uri: str = "mongodb://localhost:27017"):
        self.mongo_uri = mongo_uri
        self.client = None
        self.db = None
        self.collection = None
        
    def connect(self, db_name: str, collection_name: str):
        """Connect to MongoDB database and collection"""
        try:
            self.client = MongoClient(self.mongo_uri)
            self.db = self.client[db_name]
            self.collection = self.db[collection_name]
            print(f"Connected to {db_name}.{collection_name}")
            return True
        except Exception as e:
            print(f"Error connecting to MongoDB: {e}")
            return False
    
    def find_statute_sections(self, statute_name: str) -> List[Dict]:
        """Find sections for a given statute name"""
        try:
            print(f"Searching for statute: '{statute_name}'")
            print(f"Database: {self.db.name}, Collection: {self.collection.name}")
            
            # Multiple search strategies
            search_strategies = [
                # Strategy 1: Exact match in Statute_Name
                {"Statute_Name": statute_name},
                # Strategy 2: Case-insensitive regex match
                {"Statute_Name": {"$regex": re.escape(statute_name), "$options": "i"}},
                # Strategy 3: Lenient partial word match
                {"Statute_Name": {"$regex": f".*{re.escape(statute_name)}.*", "$options": "i"}},
            ]
            
            for i, strategy in enumerate(search_strategies):
                print(f"Trying strategy {i+1}: {strategy}")
                result = self.collection.find_one(strategy)
                if result and result.get('Sections'):
                    print(f"Found statute with {len(result['Sections'])} sections")
                    return result['Sections']
                elif result:
                    print(f"Found statute but no sections")
                else:
                    print(f"No statute found with strategy {i+1}")
            
            # If no exact match, try fuzzy matching
            print("Trying fuzzy matching...")
            all_statutes = list(self.collection.find({}, {"Statute_Name": 1, "Sections": 1}))
            print(f"Total statutes in collection: {len(all_statutes)}")
            
            if all_statutes:
                statute_names = [s.get('Statute_Name', '') for s in all_statutes]
                matches = process.extract(statute_name, statute_names, scorer=fuzz.WRatio, limit=3)
                
                print(f"Top 3 fuzzy matches:")
                for match, score in matches:
                    print(f"  '{match}' (score: {score:.2f})")
                
                if matches and matches[0][1] > 70:  # High confidence threshold
                    best_match_name = matches[0][0]
                    for statute in all_statutes:
                        if statute.get('Statute_Name') == best_match_name:
                            print(f"Using fuzzy match: '{best_match_name}'")
                            return statute.get('Sections', [])
            
            print("No statute found with any strategy")
            return []
            
        except Exception as e:
            print(f"Error finding statute sections: {e}")
            return []

class AIDateExtractor:
    """Handles AI-powered date extraction"""
    
    def __init__(self, api_key: str, azure_endpoint: str, model: str = "gpt-4o", api_version: str = "2024-11-01-preview"):
        """
        Initialize Azure OpenAI client
        
        Args:
            api_key: Azure OpenAI API key
            azure_endpoint: Azure OpenAI endpoint URL (e.g., https://your-resource.openai.azure.com/)
            model: GPT model to use (default: gpt-4o)
            api_version: Azure OpenAI API version (default: 2024-11-01-preview)
        """
        self.client = AzureOpenAI(
            api_key=api_key,
            api_version=api_version,
            azure_endpoint=azure_endpoint
        )
        self.model = model
        
        # Initialize GPT optimization utilities if available
        if GPT_UTILS_AVAILABLE:
            self.cache = gpt_cache
            self.rate_limiter = AdvancedRateLimiter()
            self.prompt_optimizer = PromptOptimizer()
            self.monitor = gpt_monitor
        else:
            self.cache = None
            self.rate_limiter = None
            self.prompt_optimizer = None
            self.monitor = None
    
    def extract_dates_from_sections(self, statute_name: str, sections: List[Dict]) -> Tuple[str, List[str], float, str]:
        """Extract dates from statute sections using AI"""
        try:
            print(f"\n🔍 Processing statute: {statute_name}")
            print(f"📄 Found {len(sections)} sections")
            
            # Sort sections to prioritize important ones
            sorted_sections = self.sort_sections(sections)
            print(f"📋 Top 3 sections after sorting: {[s.get('Section', '') for s in sorted_sections[:3]]}")
            
            # Prepare context for AI
            context = self.prepare_context(statute_name, sorted_sections)
            print(f"📝 Context length: {len(context)} characters")
            
            # Create AI prompt
            prompt = self.create_extraction_prompt(statute_name, context)
            
            # Call AI API
            response = self.call_ai_api(prompt)
            print(f"🤖 AI Response: {response[:200]}..." if response else "❌ No AI response")
            
            # Parse AI response
            extracted_date, all_dates, confidence, reasoning = self.parse_ai_response(response)
            print(f"📅 Extracted date: {extracted_date}")
            print(f"📅 All dates: {all_dates}")
            print(f"🎯 Confidence: {confidence}")
            print(f"💭 Reasoning: {reasoning}")
            
            return extracted_date, all_dates, confidence, reasoning
            
        except Exception as e:
            print(f"❌ Error extracting dates: {e}")
            return "", [], 0.0, f"Error: {str(e)}"
    
    def sort_sections(self, sections: List[Dict]) -> List[Dict]:
        """Sort sections to prioritize important ones for date extraction"""
        if not sections:
            return sections
        
        # Define priority order for better date extraction
        # Based on typical statute structure and likelihood of containing dates
        priority_order = [
            'preamble',           # Often contains enactment/promulgation info
            'title',              # May contain short title with date
            'short title',        # Often includes year
            'commencement',       # Directly about when law comes into force
            '1',                  # Section 1 often has key dates
            '2',                  # Section 2
            '3',                  # Section 3
            '4',                  # Section 4
            '5',                  # Section 5
            'rules',              # Rules may contain dates
            'regulations',        # Regulations may contain dates
            'schedule',           # Schedules may have dates
            'appendix',           # Appendices may have dates
        ]
        
        def get_priority(section):
            section_name = section.get('Section', '').lower().strip()
            
            # Direct match for priority items
            for i, priority_item in enumerate(priority_order):
                if section_name == priority_item:
                    return i
            
            # Check if section name contains priority keywords
            for i, priority_item in enumerate(priority_order):
                if priority_item in section_name:
                    return i
            
            # Try to extract pure numbers for numeric sorting
            import re
            if section_name.isdigit():
                return 50 + int(section_name)  # Pure numbers get medium priority
            
            # Try to extract section numbers from text like "section 10"
            number_match = re.search(r'section\s*(\d+)', section_name, re.IGNORECASE)
            if number_match:
                return 100 + int(number_match.group(1))
            
            # Try to extract any number for sorting
            number_match = re.search(r'(\d+)', section_name)
            if number_match:
                return 200 + int(number_match.group(1))
            
            return 1000  # Lower priority for others
        
        return sorted(sections, key=get_priority)
    
    def prepare_context(self, statute_name: str, sections: List[Dict]) -> str:
        """Prepare context from sections for AI analysis"""
        context_parts = [f"Statute Name: {statute_name}\n\n"]
        
        # Include first few sections (limit to avoid token limits)
        for i, section in enumerate(sections[:5]):  # Limit to first 5 sections
            section_name = section.get('Section', '')
            section_text = section.get('Statute', '')
            
            if section_text:
                context_parts.append(f"Section: {section_name}\n{section_text}\n")
        
        return "\n".join(context_parts)
    
    def create_extraction_prompt(self, statute_name: str, context: str) -> str:
        """Create AI prompt for date extraction"""
        return f"""You are an expert legal document analyst. Your task is to extract the PROMULGATION/PUBLISHING DATE from the following statute text.

IMPORTANT: Focus specifically on the date when the statute was promulgated/published in the official gazette, NOT enactment dates or other dates.

Statute: {statute_name}

Context:
{context}

Instructions:
1. Look for dates that indicate when the statute was promulgated, published, or came into force
2. Focus on official publication dates in gazettes, especially dates mentioned in:
   - Gazette notifications from any Pakistani province, federal level, or historical territories:
     * "Gazette of Pakistan" / "Pakistan Gazette"
     * "Gazette of Punjab" / "Punjab Gazette" / "Gazette of Punjab Extraordinary"
     * "Gazette of Sindh" / "Sindh Gazette" / "Gazette of Sindh Extraordinary"
     * "Gazette of Khyber Pakhtunkhwa" / "KP Gazette" / "Gazette of KPK Extraordinary"
     * "Gazette of Balochistan" / "Balochistan Gazette" / "Gazette of Balochistan Extraordinary"
     * "Gazette of Gilgit-Baltistan" / "GB Gazette"
     * "Gazette of Azad Jammu and Kashmir" / "AJK Gazette"
     * "Gazette of East Pakistan" / "East Pakistan Gazette"
     * "Gazette of West Pakistan" / "West Pakistan Gazette"
   - Official notification numbers with dates (e.g., "dated 4-3-2016", "No. PAP/Legis-2(99)/2015/1389")
   - Governor assent dates (e.g., "assented to by the Governor on March 3, 2016")
   - President assent dates (e.g., "assented to by the President on...")
   - Provincial Assembly passage dates (e.g., "passed by the Provincial Assembly on...")
3. Ignore dates that are not related to promulgation/publishing
4. If multiple dates are found, identify the most relevant one as the primary date
5. Convert all dates to DD-MMM-YYYY format (e.g., "4th March, 2016" becomes "04-Mar-2016")

Please respond in the following JSON format:
{{
    "primary_date": "DD-MMM-YYYY format (e.g., 04-Mar-2016)",
    "all_dates_found": ["DD-MMM-YYYY", "DD-MMM-YYYY", "DD-MMM-YYYY"],
    "confidence": 0.95,
    "reasoning": "Brief explanation of why this date was selected"
}}

If no relevant date is found, respond with:
{{
    "primary_date": "",
    "all_dates_found": [],
    "confidence": 0.0,
    "reasoning": "No promulgation/publishing date found"
}}"""
    
    def call_ai_api(self, prompt: str) -> str:
        """Call AI API with rate limiting and caching"""
        try:
            # Use cache if available
            if self.cache:
                cached_response = self.cache.get(prompt)
                if cached_response:
                    # Log cache hit
                    if self.monitor:
                        self.monitor.log_call("cached", success=True)
                    return cached_response
            
            # Use rate limiter if available
            if self.rate_limiter:
                # Check if we can make a request, if not, wait
                if not self.rate_limiter.rate_limiter.can_make_request():
                    wait_time = self.rate_limiter.rate_limiter.get_wait_time()
                    if wait_time > 0:
                        import time
                        time.sleep(wait_time)
                        # Log rate limited call
                        if self.monitor:
                            self.monitor.log_call("rate_limited", success=True)
            
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                max_tokens=500,
                temperature=0.1
            )
            
            result = response.choices[0].message.content
            
            # Cache result if available
            if self.cache:
                self.cache.set(prompt, result)
            
            # Record successful request in rate limiter
            if self.rate_limiter:
                self.rate_limiter.rate_limiter.record_request()
            
            # Monitor usage if available
            if self.monitor:
                self.monitor.log_call("api", success=True)
            
            return result
            
        except Exception as e:
            print(f"AI API Error: {e}")
            
            # Record failure in rate limiter if available
            if self.rate_limiter:
                self.rate_limiter.circuit_breaker.on_failure(e)
            
            # Log error in monitor
            if self.monitor:
                self.monitor.log_call("api", success=False, error=str(e))
            
            return ""
    
    def parse_ai_response(self, response: str) -> Tuple[str, List[str], float, str]:
        """Parse AI response to extract dates and metadata"""
        try:
            print(f"🔍 Parsing AI response: {response[:100]}...")
            
            # Try to extract JSON from response
            json_match = re.search(r'\{.*\}', response, re.DOTALL)
            if json_match:
                json_str = json_match.group()
                print(f"📋 Found JSON: {json_str}")
                data = json.loads(json_str)
                
                primary_date = data.get('primary_date', '')
                all_dates = data.get('all_dates_found', [])
                confidence = data.get('confidence', 0.0)
                reasoning = data.get('reasoning', '')
                
                print(f"📅 Raw primary_date: {primary_date}")
                print(f"📅 Raw all_dates: {all_dates}")
                
                # Format dates to DD-MMM-YYYY
                primary_date = format_date_to_dd_mmm_yyyy(primary_date)
                all_dates = [format_date_to_dd_mmm_yyyy(date) for date in all_dates if date.strip()]
                
                print(f"📅 Formatted primary_date: {primary_date}")
                print(f"📅 Formatted all_dates: {all_dates}")
                
                return primary_date, all_dates, confidence, reasoning
            
            # Fallback: try to extract dates using regex
            print("🔄 JSON parsing failed, trying regex fallback...")
            date_patterns = [
                r'\b\d{1,2}-\w{3}-\d{4}\b',  # DD-MMM-YYYY
                r'\b\d{4}-\d{2}-\d{2}\b',    # YYYY-MM-DD
                r'\b\d{1,2}/\d{1,2}/\d{4}\b', # MM/DD/YYYY
                r'\b\d{1,2}-\d{1,2}-\d{4}\b', # DD-M-YYYY (e.g., 4-3-2016)
                r'\b\d{1,2}st\s+\w+\s*,\s*\d{4}\b',  # 1st March, 2016
                r'\b\d{1,2}nd\s+\w+\s*,\s*\d{4}\b',  # 2nd March, 2016
                r'\b\d{1,2}rd\s+\w+\s*,\s*\d{4}\b',  # 3rd March, 2016
                r'\b\d{1,2}th\s+\w+\s*,\s*\d{4}\b',  # 4th March, 2016
                r'\b\w+\s+\d{1,2}\s*,\s*\d{4}\b',    # March 3, 2016
            ]
            
            all_dates = []
            for i, pattern in enumerate(date_patterns):
                dates = re.findall(pattern, response)
                if dates:
                    print(f"🔍 Pattern {i+1} found dates: {dates}")
                all_dates.extend(dates)
            
            print(f"📅 Total dates found by regex: {all_dates}")
            
            if all_dates:
                # Format all found dates
                formatted_dates = [format_date_to_dd_mmm_yyyy(date) for date in all_dates]
                print(f"📅 Formatted dates: {formatted_dates}")
                return formatted_dates[0], formatted_dates, 0.5, "Extracted using regex fallback"
            
            print("❌ No dates found by regex either")
            return "", [], 0.0, "No dates found"
            
        except Exception as e:
            print(f"Error parsing AI response: {e}")
            return "", [], 0.0, f"Parse error: {str(e)}"

class DateExtractorGUI:
    """GUI for reviewing and approving extracted dates"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("AI Date Extractor")
        self.root.geometry("1400x900")
        
        # Data
        self.excel_processor = None
        self.db_connector = None
        self.ai_extractor = None
        self.statutes = []
        self.current_index = 0
        
        # Configuration
        self.config = self.load_config()
        
        self.init_ui()
    
    def load_config(self) -> Dict:
        """Load configuration from file"""
        config_file = os.path.join(os.path.dirname(__file__), "config_ai_extractor.json")
        try:
            with open(config_file, 'r') as f:
                return json.load(f)
        except:
            return {
                "mongo_uri": "mongodb://localhost:27017",
                "db_name": "Statute",
                "collection_name": "normalized_statutes",
                "api_key": "",
                "azure_endpoint": "",
                "model": "gpt-4o",
                "api_version": "2024-11-01-preview"
            }
    
    def init_ui(self):
        """Initialize the GUI"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Top controls
        self.create_controls(main_frame)
        
        # Left panel - Statute list
        self.create_statute_list(main_frame)
        
        # Right panel - Details and editing
        self.create_details_panel(main_frame)
        
        # Bottom panel - Actions
        self.create_actions_panel(main_frame)
    
    def create_controls(self, parent):
        """Create top control panel"""
        controls_frame = ttk.LabelFrame(parent, text="Controls", padding="5")
        controls_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Excel file selection
        ttk.Label(controls_frame, text="Excel File:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.excel_path_var = tk.StringVar()
        ttk.Entry(controls_frame, textvariable=self.excel_path_var, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        ttk.Button(controls_frame, text="Browse", command=self.browse_excel).grid(row=0, column=2, padx=(0, 10))
        
        # Load button
        ttk.Button(controls_frame, text="Load Data", command=self.load_data).grid(row=0, column=3, padx=(0, 10))
        
        # Process buttons
        ttk.Button(controls_frame, text="Process Current", command=self.process_with_ai).grid(row=0, column=4, padx=(0, 10))
        ttk.Button(controls_frame, text="Process All", command=self.process_all_statutes).grid(row=0, column=5, padx=(0, 10))
        
        # Save button
        ttk.Button(controls_frame, text="Save to Excel", command=self.save_to_excel).grid(row=0, column=6)
        
        # GPT Stats button
        ttk.Button(controls_frame, text="GPT Stats", command=self.show_gpt_stats).grid(row=0, column=7, padx=(0, 10))
        
        # Status
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(controls_frame, textvariable=self.status_var, foreground="blue").grid(row=1, column=0, columnspan=8, sticky=tk.W, pady=(5, 0))
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(controls_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=2, column=0, columnspan=8, sticky=(tk.W, tk.E), pady=(5, 0))
    
    def create_statute_list(self, parent):
        """Create statute list panel"""
        list_frame = ttk.LabelFrame(parent, text="Statutes with Missing Dates", padding="5")
        list_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))
        list_frame.columnconfigure(0, weight=1)
        list_frame.rowconfigure(1, weight=1)
        
        # Search
        ttk.Label(list_frame, text="Search:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.on_search_change)
        ttk.Entry(list_frame, textvariable=self.search_var).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Statute list
        self.statute_listbox = tk.Listbox(list_frame, height=20)
        self.statute_listbox.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.statute_listbox.yview)
        self.statute_listbox.configure(yscrollcommand=scrollbar.set)
        scrollbar.grid(row=1, column=2, sticky=(tk.N, tk.S))
        
        self.statute_listbox.bind('<<ListboxSelect>>', self.on_statute_select)
    
    def create_details_panel(self, parent):
        """Create details and editing panel"""
        details_frame = ttk.LabelFrame(parent, text="Details & Editing", padding="5")
        details_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S))
        details_frame.columnconfigure(0, weight=1)
        details_frame.rowconfigure(6, weight=1)
        
        # Statute name
        ttk.Label(details_frame, text="Statute Name:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        self.statute_name_var = tk.StringVar()
        ttk.Entry(details_frame, textvariable=self.statute_name_var, width=50).grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Best Date
        ttk.Label(details_frame, text="Best Date:").grid(row=1, column=0, sticky=tk.W, pady=(0, 5))
        self.best_date_var = tk.StringVar()
        ttk.Entry(details_frame, textvariable=self.best_date_var, width=20).grid(row=1, column=1, sticky=tk.W, pady=(0, 5))
        
        # All Dates Extracted
        ttk.Label(details_frame, text="All Dates Extracted:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        self.all_dates_var = tk.StringVar()
        ttk.Entry(details_frame, textvariable=self.all_dates_var, width=50).grid(row=2, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Selection Reason
        ttk.Label(details_frame, text="Selection Reason:").grid(row=3, column=0, sticky=tk.W, pady=(0, 5))
        self.selection_reason_var = tk.StringVar()
        ttk.Entry(details_frame, textvariable=self.selection_reason_var, width=50).grid(row=3, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Search Method
        ttk.Label(details_frame, text="Search Method:").grid(row=4, column=0, sticky=tk.W, pady=(0, 5))
        self.search_method_var = tk.StringVar()
        ttk.Entry(details_frame, textvariable=self.search_method_var, width=50).grid(row=4, column=1, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Confidence
        ttk.Label(details_frame, text="Confidence:").grid(row=5, column=0, sticky=tk.W, pady=(0, 5))
        self.confidence_var = tk.StringVar()
        ttk.Label(details_frame, textvariable=self.confidence_var).grid(row=5, column=1, sticky=tk.W, pady=(0, 5))
        
        # Sections used
        ttk.Label(details_frame, text="Sections Used:").grid(row=6, column=0, sticky=tk.W, pady=(0, 5))
        self.sections_text = scrolledtext.ScrolledText(details_frame, height=10, wrap=tk.WORD)
        self.sections_text.grid(row=7, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 5))
    
    def create_actions_panel(self, parent):
        """Create bottom actions panel"""
        actions_frame = ttk.Frame(parent)
        actions_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Button(actions_frame, text="Approve", command=self.approve_date).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(actions_frame, text="Approve All", command=self.approve_all_dates).grid(row=0, column=1, padx=(0, 10))
        ttk.Button(actions_frame, text="Reject", command=self.reject_date).grid(row=0, column=2, padx=(0, 10))
        ttk.Button(actions_frame, text="Skip", command=self.skip_statute).grid(row=0, column=3, padx=(0, 10))
        ttk.Button(actions_frame, text="Previous", command=self.previous_statute).grid(row=0, column=4, padx=(0, 10))
        ttk.Button(actions_frame, text="Next", command=self.next_statute).grid(row=0, column=5, padx=(0, 10))
    
    def browse_excel(self):
        """Browse for Excel file"""
        filename = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.excel_path_var.set(filename)
    
    def load_data(self):
        """Load Excel data and find statutes with missing dates"""
        def load_thread():
            try:
                self.status_var.set("Loading Excel data...")
                self.progress_var.set(10)
                
                excel_path = self.excel_path_var.get()
                if not excel_path or not os.path.exists(excel_path):
                    messagebox.showerror("Error", "Please select a valid Excel file")
                    return
                
                # Load Excel
                self.excel_processor = ExcelProcessor(excel_path)
                if not self.excel_processor.load_workbook():
                    messagebox.showerror("Error", "Failed to load Excel file")
                    return
                
                self.progress_var.set(30)
                self.status_var.set("Finding statutes with missing dates...")
                
                # Get statutes with missing dates
                self.statutes = self.excel_processor.get_statutes_with_missing_dates()
                
                self.progress_var.set(50)
                self.status_var.set("Connecting to database...")
                
                # Connect to database
                self.db_connector = DatabaseConnector(self.config.get("mongo_uri"))
                if not self.db_connector.connect(
                    self.config.get("db_name"),
                    self.config.get("collection_name")
                ):
                    messagebox.showerror("Error", "Failed to connect to database")
                    return
                
                self.progress_var.set(70)
                self.status_var.set("Initializing AI...")
                
                # Initialize AI
                api_key = self.config.get("api_key")
                azure_endpoint = self.config.get("azure_endpoint")
                model = self.config.get("model", "gpt-4o")
                api_version = self.config.get("api_version", "2024-11-01-preview")
                if not api_key or not azure_endpoint:
                    messagebox.showerror("Error", "API key and Azure endpoint URL required in config")
                    return
                
                self.ai_extractor = AIDateExtractor(api_key, azure_endpoint, model, api_version)
                
                self.progress_var.set(100)
                self.status_var.set(f"Loaded {len(self.statutes)} statutes with missing dates")
                
                # Update UI
                self.root.after(0, self.update_statute_list)
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Failed to load data: {e}"))
                self.root.after(0, lambda: self.status_var.set("Error loading data"))
        
        threading.Thread(target=load_thread, daemon=True).start()
    
    def update_statute_list(self):
        """Update the statute listbox"""
        self.statute_listbox.delete(0, tk.END)
        for statute in self.statutes:
            self.statute_listbox.insert(tk.END, statute.statute_name)
        
        if self.statutes:
            self.statute_listbox.selection_set(0)
            self.current_index = 0
            self.show_statute_details()
    
    def on_search_change(self, *args):
        """Handle search text change"""
        search_text = self.search_var.get().lower()
        
        self.statute_listbox.delete(0, tk.END)
        for i, statute in enumerate(self.statutes):
            if search_text in statute.statute_name.lower():
                self.statute_listbox.insert(tk.END, statute.statute_name)
                if i == self.current_index:
                    self.statute_listbox.selection_set(tk.END)
    
    def on_statute_select(self, event):
        """Handle statute selection"""
        selection = self.statute_listbox.curselection()
        if selection:
            self.current_index = selection[0]
            self.show_statute_details()
    
    def show_statute_details(self):
        """Show details for current statute"""
        if not self.statutes or self.current_index >= len(self.statutes):
            return
        
        statute = self.statutes[self.current_index]
        
        self.statute_name_var.set(statute.statute_name)
        self.best_date_var.set(statute.extracted_date)
        self.all_dates_var.set(statute.all_dates_extracted)
        self.selection_reason_var.set(statute.selection_reason)
        self.search_method_var.set(statute.search_method)
        self.confidence_var.set(f"{statute.confidence:.2f}")
        
        # Show sections used
        self.sections_text.delete(1.0, tk.END)
        if statute.sections_used:
            self.sections_text.insert(1.0, "\n".join(statute.sections_used))
    
    def process_with_ai(self):
        """Process current statute with AI"""
        if not self.statutes or self.current_index >= len(self.statutes):
            messagebox.showwarning("Warning", "No statute selected")
            return
        
        def process_thread():
            try:
                statute = self.statutes[self.current_index]
                self.status_var.set(f"Processing {statute.statute_name}...")
                
                # Find sections in database
                sections = self.db_connector.find_statute_sections(statute.statute_name)
                if not sections:
                    self.root.after(0, lambda: messagebox.showwarning("Warning", f"No sections found for {statute.statute_name}"))
                    return
                
                # Extract dates using AI
                primary_date, all_dates, confidence, reasoning = self.ai_extractor.extract_dates_from_sections(
                    statute.statute_name, sections
                )
                
                # Update statute record
                statute.extracted_date = primary_date
                statute.all_dates_extracted = ", ".join(all_dates) if all_dates else primary_date
                statute.selection_reason = reasoning
                statute.search_method = "AI GPT-4o Analysis"
                statute.confidence = confidence
                statute.sections_used = [s.get('Section', '') for s in sections[:3]]  # First 3 sections
                
                self.root.after(0, self.show_statute_details)
                self.root.after(0, lambda: self.status_var.set("AI processing completed"))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"AI processing failed: {e}"))
                self.root.after(0, lambda: self.status_var.set("AI processing failed"))
        
        threading.Thread(target=process_thread, daemon=True).start()
    
    def process_all_statutes(self):
        """Process all statutes with AI in batch"""
        if not self.statutes:
            messagebox.showwarning("Warning", "No statutes loaded")
            return
        
        # Confirm with user
        result = messagebox.askyesno(
            "Confirm Batch Processing", 
            f"Process all {len(self.statutes)} statutes with AI?\n\nThis may take several minutes."
        )
        if not result:
            return
        
        def batch_process_thread():
            try:
                total_statutes = len(self.statutes)
                processed_count = 0
                success_count = 0
                error_count = 0
                dates_found_count = 0
                dates_not_found_count = 0
                
                self.status_var.set(f"Starting batch processing of {total_statutes} statutes...")
                self.progress_var.set(0)
                
                for i, statute in enumerate(self.statutes):
                    try:
                        # Update progress
                        progress = (i / total_statutes) * 100
                        self.progress_var.set(progress)
                        self.status_var.set(f"Processing {i+1}/{total_statutes}: {statute.statute_name}")
                        
                        # Find sections in database
                        sections = self.db_connector.find_statute_sections(statute.statute_name)
                        if not sections:
                            print(f"No sections found for {statute.statute_name}")
                            error_count += 1
                            continue
                        
                        # Extract dates using AI
                        primary_date, all_dates, confidence, reasoning = self.ai_extractor.extract_dates_from_sections(
                            statute.statute_name, sections
                        )
                        
                        # Update statute record
                        statute.extracted_date = primary_date
                        statute.all_dates_extracted = ", ".join(all_dates) if all_dates else primary_date
                        statute.selection_reason = reasoning
                        statute.search_method = "AI GPT-4o Analysis"
                        statute.confidence = confidence
                        statute.sections_used = [s.get('Section', '') for s in sections[:3]]  # First 3 sections
                        
                        # Count dates found vs not found
                        if primary_date and primary_date.strip():
                            dates_found_count += 1
                        else:
                            dates_not_found_count += 1
                        
                        success_count += 1
                        
                        # Small delay to avoid overwhelming the API
                        import time
                        time.sleep(0.5)
                        
                    except Exception as e:
                        print(f"Error processing {statute.statute_name}: {e}")
                        error_count += 1
                    
                    processed_count += 1
                
                # Final progress update
                self.progress_var.set(100)
                self.status_var.set(f"Batch processing completed: {success_count} success, {error_count} errors")
                
                # Update UI
                self.root.after(0, self.update_statute_list)
                self.root.after(0, lambda: messagebox.showinfo(
                    "Batch Processing Complete", 
                    f"Processed {processed_count} statutes:\n\n"
                    f"📊 Processing Results:\n"
                    f"✅ Successfully Processed: {success_count}\n"
                    f"❌ Processing Errors: {error_count}\n\n"
                    f"📅 Date Extraction Results:\n"
                    f"🎯 Dates Found: {dates_found_count}\n"
                    f"❓ Dates Not Found: {dates_not_found_count}\n\n"
                    f"📝 Note: Statutes with 'Dates Not Found' may need manual review."
                ))
                
            except Exception as e:
                self.root.after(0, lambda: messagebox.showerror("Error", f"Batch processing failed: {e}"))
                self.root.after(0, lambda: self.status_var.set("Batch processing failed"))
        
        threading.Thread(target=batch_process_thread, daemon=True).start()
    
    def show_gpt_stats(self):
        """Show GPT optimization statistics"""
        if not self.ai_extractor:
            messagebox.showinfo("GPT Stats", "AI extractor not initialized. Please load data first.")
            return
            
        if not self.ai_extractor.monitor:
            messagebox.showinfo("GPT Stats", 
                "GPT monitoring not available.\n\n"
                "This could be because:\n"
                "• GPT utilities failed to import\n"
                "• Check console for import errors\n"
                "• Ensure utils folder is accessible\n\n"
                "Stats will be available after processing if monitoring is working.")
            return
        
        try:
            stats = self.ai_extractor.monitor.get_optimization_stats()
            
            stats_text = f"""GPT Optimization Statistics:

📊 Total Calls: {stats.get('total_calls', 0)}
🎯 Cache Hit Rate: {stats.get('cache_hit_rate_percent', 0):.1f}%
🔄 Rate Limited Calls: {stats.get('rate_limited_calls', 0)}
📈 API Call Rate: {stats.get('api_call_rate_percent', 0):.1f}%
❌ Error Rate: {stats.get('error_rate_percent', 0):.1f}%
💰 Estimated Cost Saved: ${stats.get('estimated_cost_saved_usd', 0):.2f}
⏱️  Uptime: {stats.get('uptime_hours', 0):.1f} hours

🔧 Optimization Status:
• Cache: {'✅ Enabled' if self.ai_extractor.cache else '❌ Disabled'}
• Rate Limiter: {'✅ Enabled' if self.ai_extractor.rate_limiter else '❌ Disabled'}
• Prompt Optimizer: {'✅ Enabled' if self.ai_extractor.prompt_optimizer else '❌ Disabled'}
• Monitor: {'✅ Enabled' if self.ai_extractor.monitor else '❌ Disabled'}"""
            
            # Create a new window to display stats
            stats_window = tk.Toplevel(self.root)
            stats_window.title("GPT Optimization Statistics")
            stats_window.geometry("600x500")
            
            text_widget = scrolledtext.ScrolledText(stats_window, wrap=tk.WORD, padx=10, pady=10)
            text_widget.pack(fill=tk.BOTH, expand=True)
            text_widget.insert(tk.END, stats_text)
            text_widget.config(state=tk.DISABLED)
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to get GPT stats: {e}")
    
    def approve_date(self):
        """Approve the extracted date"""
        if not self.statutes or self.current_index >= len(self.statutes):
            return
        
        statute = self.statutes[self.current_index]
        
        # Update with current values from UI
        statute.extracted_date = self.best_date_var.get()
        statute.all_dates_extracted = self.all_dates_var.get()
        statute.selection_reason = self.selection_reason_var.get()
        statute.search_method = self.search_method_var.get()
        
        # Save to Excel
        self.excel_processor.save_to_excel(statute)
        
        # Move to next statute
        self.next_statute()
        
        messagebox.showinfo("Success", "Date approved and saved")
    
    def approve_all_dates(self):
        """Approve all processed dates and save to Excel"""
        if not self.statutes:
            messagebox.showwarning("Warning", "No statutes loaded")
            return
        
        # Count how many have extracted dates
        statutes_with_dates = [s for s in self.statutes if s.extracted_date and s.extracted_date.strip()]
        statutes_without_dates = [s for s in self.statutes if not s.extracted_date or not s.extracted_date.strip()]
        
        if not statutes_with_dates:
            messagebox.showwarning("Warning", "No statutes have extracted dates to approve")
            return
        
        # Confirm with user
        result = messagebox.askyesno(
            "Confirm Approve All", 
            f"Approve and save all {len(statutes_with_dates)} statutes with extracted dates?\n\n"
            f"📊 Summary:\n"
            f"✅ Statutes with dates: {len(statutes_with_dates)}\n"
            f"❓ Statutes without dates: {len(statutes_without_dates)}\n\n"
            f"This will save all approved dates to Excel."
        )
        if not result:
            return
        
        try:
            # Save all statutes with dates to Excel
            saved_count = 0
            for statute in statutes_with_dates:
                self.excel_processor.save_to_excel(statute)
                saved_count += 1
            
            # Save workbook
            if self.excel_processor.save_workbook():
                messagebox.showinfo(
                    "Success", 
                    f"Approved and saved {saved_count} statutes to Excel!\n\n"
                    f"📊 Results:\n"
                    f"✅ Saved: {saved_count}\n"
                    f"❓ Skipped (no dates): {len(statutes_without_dates)}\n\n"
                    f"All changes have been saved to the Excel file."
                )
                self.status_var.set(f"Approved and saved {saved_count} statutes")
            else:
                messagebox.showerror("Error", "Failed to save Excel file")
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to approve all dates: {e}")
    
    def reject_date(self):
        """Reject the extracted date"""
        if not self.statutes or self.current_index >= len(self.statutes):
            return
        
        statute = self.statutes[self.current_index]
        
        # Clear extracted data
        statute.extracted_date = ""
        statute.all_dates_extracted = ""
        statute.selection_reason = "Rejected by user"
        statute.search_method = "Manual review"
        
        self.show_statute_details()
    
    def skip_statute(self):
        """Skip current statute"""
        self.next_statute()
    
    def previous_statute(self):
        """Go to previous statute"""
        if self.current_index > 0:
            self.current_index -= 1
            self.statute_listbox.selection_clear(0, tk.END)
            self.statute_listbox.selection_set(self.current_index)
            self.show_statute_details()
    
    def next_statute(self):
        """Go to next statute"""
        if self.current_index < len(self.statutes) - 1:
            self.current_index += 1
            self.statute_listbox.selection_clear(0, tk.END)
            self.statute_listbox.selection_set(self.current_index)
            self.show_statute_details()
    
    def save_to_excel(self):
        """Save all changes to Excel"""
        try:
            if self.excel_processor.save_workbook():
                messagebox.showinfo("Success", "All changes saved to Excel")
                self.status_var.set("Changes saved successfully")
            else:
                messagebox.showerror("Error", "Failed to save changes")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save: {e}")

def main():
    root = tk.Tk()
    app = DateExtractorGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()