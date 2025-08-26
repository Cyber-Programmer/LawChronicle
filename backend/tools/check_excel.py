#!/usr/bin/env python3
"""
Check the content of the generated Excel file
"""

import openpyxl

def check_excel_content():
    """Check what's in the Excel file"""
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook('ai_search_results.xlsx')
        
        print("📊 Excel File Analysis")
        print("=" * 40)
        print(f"📋 Worksheets: {wb.sheetnames}")
        
        # Check each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n📄 Sheet: {sheet_name}")
            print(f"   Max row: {ws.max_row}")
            print(f"   Max col: {ws.max_column}")
            
            # Show first few rows
            if ws.max_row > 0:
                print(f"   Content preview:")
                for row in range(1, min(6, ws.max_row + 1)):
                    row_data = []
                    for col in range(1, min(6, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    print(f"     Row {row}: {row_data}")
            else:
                print("   ❌ Sheet is empty")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error reading Excel: {e}")

if __name__ == "__main__":
    check_excel_content()
